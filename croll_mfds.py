import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import datetime
import pandas as pd
import re
import json
import pyperclip  # 클립보드 처리를 위한 라이브러리 추가
import time  # 딜레이를 위한 time 모듈 추가
import subprocess
import os

# 공통 헤더
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.61 Safari/537.36'
}

# 식약처 관련 URL 목록
food_drug_urls = [
    {'url': 'https://www.mfds.go.kr/brd/m_824/list.do', '출처': '서울지방 식약청'},
    {'url': 'https://www.mfds.go.kr/brd/m_841/list.do', '출처': '경인지방 식약청'},
    {'url': 'https://www.mfds.go.kr/brd/m_765/list.do', '출처': '부산지방 식약청'},
    {'url': 'https://www.mfds.go.kr/brd/m_891/list.do', '출처': '대전지방 식약청'},
    {'url': 'https://www.mfds.go.kr/brd/m_875/list.do', '출처': '광주지방 식약청'},
    {'url': 'https://www.mfds.go.kr/brd/m_857/list.do', '출처': '대구지방 식약청'},
    {'url': 'https://kcia.or.kr/home/notice/notice.php?page=1', '출처': '대한화장품협회'},
    {'url': 'https://www.mfds.go.kr/brd/m_1059/list.do', '출처': '식약처 공무원지침서'},
    {'url': 'https://www.mfds.go.kr/brd/m_1060/list.do', '출처': '식약처 민원인안내서'}
]

# 게시물을 저장할 리스트
all_posts = []

def get_detail_content(url):
    """
    게시글의 상세 내용을 크롤링하는 함수
    """
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 식약청 지방청의 경우
        if "mfds.go.kr" in url:
            content_div = soup.find('div', class_='bv_cont')
        # 대한화장품협회의 경우
        elif "kcia.or.kr" in url:
            content_div = soup.find('div', class_='view_area')
        else:
            return ""
            
        if content_div:
            # 모든 텍스트 추출 (줄바꿈 유지)
            content = content_div.get_text(separator='\n', strip=True)
            # 연속된 빈 줄 제거
            content = re.sub(r'\n\s*\n', '\n', content)
            return content.strip()
        return ""
    except Exception as e:
        print(f"상세 내용 크롤링 중 오류 발생: {url} - {str(e)}")
        return ""

def process_mfds(url, soup, source):
    now = datetime.datetime.now()
    current_month = now.month
    last_month = (now - datetime.timedelta(days=30)).month

    for item in soup.select('ul > li'):
        title_tag = item.select_one('div.center_column > a.title')
        date_tag = item.select_one('div.right_column')
        if title_tag and date_tag:
            title = title_tag.text.strip()
            link = title_tag.get('href')
            date_str = date_tag.text.strip()
            try:
                post_date = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                if post_date.month in [current_month, last_month]:
                    if link:
                        link = urljoin(url, link)
                        # 상세 내용 크롤링 (식약청 지방청, 공무원지침서, 민원인안내서인 경우)
                        detail_content = ""
                        if any(region in source for region in ['서울', '경인', '부산', '대전', '광주', '대구', '공무원지침서', '민원인안내서']):
                            detail_content = get_detail_content(link)
                            time.sleep(0.5)  # 서버 부하 방지를 위한 딜레이
                        
                        all_posts.append({
                            '구분': source,
                            '제목': title,
                            'URL': link,
                            '등록일': date_str,
                            '상세내용': detail_content
                        })
            except ValueError:
                continue

def process_kcia(url, soup, source):
    now = datetime.datetime.now()
    current_month = now.month
    last_month = (now - datetime.timedelta(days=30)).month

    for item in soup.select('tbody > tr'):
        title_tag = item.select_one('td.left > a.link')
        date_tag = item.select('td')[-1]
        if title_tag and date_tag:
            title = title_tag.text.strip()
            link = title_tag.get('href')
            date_str = date_tag.text.strip()
            try:
                post_date = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                if post_date.month in [current_month, last_month]:
                    if link:
                        base_url = url.split('?')[0]
                        link = urljoin(base_url, link)
                        # 상세 내용 크롤링
                        detail_content = get_detail_content(link)
                        time.sleep(0.5)  # 서버 부하 방지를 위한 딜레이
                        
                        all_posts.append({
                            '구분': source,
                            '제목': title,
                            'URL': link,
                            '등록일': date_str,
                            '상세내용': detail_content
                        })
            except ValueError:
                continue

# 식약처 관련 사이트 크롤링
print("\n식약처 관련 사이트 크롤링 시작...")
for site in food_drug_urls:
    url = site['url']
    source = site['출처']
    print(f"크롤링 중: {source}")
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        if "kcia.or.kr" in url:
            process_kcia(url, soup, source)
        elif "mfds.go.kr" in url:
            process_mfds(url, soup, source)
    except Exception as e:
        print(f"Error processing {url}: {e}")

# 데이터프레임 생성 및 열 순서 지정
df = pd.DataFrame(all_posts)
df = df[['구분', '제목', '등록일', 'URL', '상세내용']]  # 상세내용 열 추가

# 등록일 기준으로 내림차순 정렬
df['등록일'] = pd.to_datetime(df['등록일'])  # 문자열을 datetime 형식으로 변환
df = df.sort_values('등록일', ascending=False)  # 내림차순 정렬
df['등록일'] = df['등록일'].dt.strftime('%Y-%m-%d')  # 다시 문자열 형식으로 변환

# 현재 날짜와 시간을 파일 이름에 포함
current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
excel_filename = f'crawling_results_{current_time}.xlsx'

try:
    # openpyxl을 사용하여 워크북과 워크시트 생성
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "크롤링 데이터"
    
    # 데이터를 워크시트에 추가
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 하이퍼링크 추가 (D열 - URL)
    for row in range(2, ws.max_row + 1):  # 헤더 제외하고 시작
        cell = ws[f'D{row}']
        if cell.value and cell.value != 'N/A':
            cell.hyperlink = cell.value
            cell.font = Font(color="0000FF", underline="single")
            cell.value = cell.value  # 표시될 텍스트는 URL 그대로
    
    # 열 너비 자동 조절
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # 최소 너비와 최대 너비 설정
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # B열(제목)은 더 넓게 설정
    ws.column_dimensions['B'].width = min(60, max(ws.column_dimensions['B'].width, 30))
    
    # E열(상세내용)도 넓게 설정
    ws.column_dimensions['E'].width = min(80, max(ws.column_dimensions['E'].width, 40))
    
    # 파일 저장
    wb.save(excel_filename)
    print(f"\n크롤링 완료!")
    print(f"총 게시글 수: {len(all_posts)}")
    print(f"크롤링 결과가 '{excel_filename}' 파일로 저장되었습니다.")
    print("- 열 너비가 자동으로 조절되었습니다.")
    print("- D열의 URL 링크는 클릭 가능한 하이퍼링크로 설정되었습니다.")
    
    # data.json 저장 (웹페이지용 고정 파일명)
    json_data = df.to_dict(orient='records')
    data_json = {
        "last_updated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "data": json_data
    }
    json_str = json.dumps(data_json, ensure_ascii=False, indent=2)
    with open('data.json', 'w', encoding='utf-8') as f:
        f.write(json_str)
    print(f"크롤링 결과가 'data.json' 파일로 저장되었습니다.")

    # 타임스탬프 JSON도 별도 저장
    json_filename = f'crawling_results_{current_time}.json'
    with open(json_filename, 'w', encoding='utf-8') as f:
        f.write(json_str)
    print(f"백업 JSON: '{json_filename}'")

    # JSON 형식으로 클립보드에 복사
    pyperclip.copy(json.dumps(json_data, ensure_ascii=False, indent=2))
    print("\n데이터가 JSON 형식으로 클립보드에 복사되었습니다.")

    # Git push (data.json을 레포지토리에 반영)
    project_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        def run_git(cmd):
            result = subprocess.run(
                cmd, cwd=project_dir, capture_output=True, text=True, timeout=30
            )
            return result

        print("\n📤 Git push 시작...")
        run_git(['git', 'add', 'data.json'])
        commit_msg = f"Update data.json ({datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})"
        commit_result = run_git(['git', 'commit', '-m', commit_msg])
        if commit_result.returncode == 0:
            push_result = run_git(['git', 'push'])
            if push_result.returncode == 0:
                print("✅ data.json이 GitHub에 push되었습니다. Cloudflare Pages 배포가 시작됩니다.")
            else:
                print(f"⚠️ Git push 실패: {push_result.stderr}")
        else:
            if 'nothing to commit' in commit_result.stdout:
                print("ℹ️ data.json에 변경사항 없음 (이전과 동일한 데이터)")
            else:
                print(f"⚠️ Git commit 실패: {commit_result.stderr}")
    except Exception as git_err:
        print(f"⚠️ Git 작업 중 오류: {git_err}")
        print("수동으로 git push 해주세요.")
    
except Exception as e:
    print(f"파일 저장 또는 클립보드 복사 중 오류 발생: {e}")
    # 오류 발생 시 기본 방식으로 저장 시도
    try:
        df.to_excel(excel_filename, index=False, engine='openpyxl')
        print(f"기본 방식으로 '{excel_filename}' 파일을 저장했습니다.")
    except Exception as e2:
        print(f"기본 저장도 실패했습니다: {e2}")